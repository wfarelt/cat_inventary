import io
import re
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import permission_required, login_required
from django.db.models import Q
from .models import Product, Category
from openpyxl import load_workbook


@permission_required('products.view_product', raise_exception=True)
def api_product_list(request):
    q = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 50))
    qs = Product.objects.select_related('category').all().order_by('code')
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(description__icontains=q))
    # filter by category (id or name)
    cat = request.GET.get('category')
    if cat:
        try:
            cid = int(cat)
            qs = qs.filter(category__id=cid)
        except Exception:
            qs = qs.filter(category__name__icontains=cat)
    # filter by has_image (1/true -> has image), else no image
    has_image = request.GET.get('has_image')
    if has_image is not None:
        if str(has_image).lower() in ('1', 'true', 'yes'):
            qs = qs.filter(image__isnull=False).exclude(image='')
        else:
            qs = qs.filter(Q(image__isnull=True) | Q(image=''))
    # price and stock range filters
    price_min = request.GET.get('price_min')
    price_max = request.GET.get('price_max')
    stock_min = request.GET.get('stock_min')
    stock_max = request.GET.get('stock_max')
    try:
        if price_min is not None and price_min != '':
            qs = qs.filter(price__gte=Decimal(price_min))
    except Exception:
        pass
    try:
        if price_max is not None and price_max != '':
            qs = qs.filter(price__lte=Decimal(price_max))
    except Exception:
        pass
    try:
        if stock_min is not None and stock_min != '':
            qs = qs.filter(stock__gte=Decimal(stock_min))
    except Exception:
        pass
    try:
        if stock_max is not None and stock_max != '':
            qs = qs.filter(stock__lte=Decimal(stock_max))
    except Exception:
        pass
    # ordering support: comma-separated fields (allowed list)
    ordering = request.GET.get('ordering')
    if ordering:
        allowed = {'code': 'code', 'description': 'description', 'price': 'price', 'stock': 'stock', 'category': 'category__name'}
        orders = []
        for part in ordering.split(','):
            p = part.strip()
            if not p:
                continue
            desc = p.startswith('-')
            key = p[1:] if desc else p
            fld = allowed.get(key)
            if fld:
                orders.append(('-' + fld) if desc else fld)
        if orders:
            qs = qs.order_by(*orders)
    # filter by category (id or name)
    cat = request.GET.get('category')
    if cat:
        try:
            cid = int(cat)
            qs = qs.filter(category__id=cid)
        except Exception:
            qs = qs.filter(category__name__icontains=cat)
    # filter by has_image (1/true -> has image), else no image
    has_image = request.GET.get('has_image')
    if has_image is not None:
        if str(has_image).lower() in ('1', 'true', 'yes'):
            qs = qs.filter(image__isnull=False).exclude(image='')
        else:
            qs = qs.filter(Q(image__isnull=True) | Q(image=''))
    total = qs.count()
    total_pages = (total + per_page - 1) // per_page if per_page > 0 else 1
    start = (page - 1) * per_page
    items = qs[start:start+per_page]
    results = []
    for p in items:
        results.append({
            'id': p.pk,
            'code': p.code,
            'description': p.description,
            'category': {'id': p.category.pk, 'name': p.category.name} if p.category else None,
            'price': str(p.price) if p.price is not None else None,
            'stock': str(p.stock) if p.stock is not None else None,
            'image': request.build_absolute_uri(p.image.url) if p.image else None,
        })
    # build next/previous urls
    def page_url(p):
        if p < 1 or p > total_pages:
            return None
        params = request.GET.copy()
        params['page'] = p
        return request.build_absolute_uri('?' + params.urlencode())

    next_url = page_url(page + 1) if page < total_pages else None
    prev_url = page_url(page - 1) if page > 1 else None

    return JsonResponse({'count': total, 'page': page, 'per_page': per_page, 'total_pages': total_pages, 'next': next_url, 'previous': prev_url, 'results': results})


@permission_required('products.view_product', raise_exception=True)
def api_product_detail(request, pk):
    p = get_object_or_404(Product.objects.select_related('category'), pk=pk)
    return JsonResponse({
        'id': p.pk,
        'code': p.code,
        'description': p.description,
        'category': {'id': p.category.pk, 'name': p.category.name} if p.category else None,
        'price': str(p.price) if p.price is not None else None,
        'stock': str(p.stock) if p.stock is not None else None,
        'image': request.build_absolute_uri(p.image.url) if p.image else None,
    })


@permission_required('products.add_product', raise_exception=True)
def api_product_import_preview(request):
    # Accepts multipart/form-data with file field 'file'
    if request.method != 'POST' or 'file' not in request.FILES:
        return HttpResponseBadRequest('file is required')
    f = request.FILES['file']
    try:
        wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return HttpResponseBadRequest(f'invalid xlsx: {e}')
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    seen_codes = set()
    code_re = re.compile(r'^[A-Za-z0-9\-_. ]+$')
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(headers, row))
        r = {k: (v if v is not None else '') for k, v in raw.items()}
        r['_row_number'] = idx
        r['errors'] = {}
        code = (r.get('code') or '').strip()
        if not code:
            r['errors'].setdefault('code', []).append('Código vacío')
        else:
            code = str(code).upper()
            r['code'] = code
            if code in seen_codes:
                r['errors'].setdefault('code', []).append('Código duplicado en archivo')
            if not code_re.match(code):
                r['errors'].setdefault('code', []).append('Código con caracteres inválidos')
            seen_codes.add(code)
        # category
        cat_name = (r.get('category') or '').strip()
        if not cat_name:
            r['errors'].setdefault('category', []).append('Categoría vacía')
        else:
            r['category'] = cat_name
        # detect existing
        rows.append(r)
    codes = [r['code'] for r in rows if r.get('code')]
    existing = set(Product.objects.filter(code__in=codes).values_list('code', flat=True))
    for r in rows:
        r['exists'] = r.get('code') in existing
    return JsonResponse({'headers': headers, 'rows': rows})
