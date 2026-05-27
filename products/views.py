import io
import zipfile
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.contrib.auth.decorators import user_passes_test
from django.http import JsonResponse
from .models import Product, Category, ProductKit, ProductKitItem
from .forms import ProductForm, CategoryForm, ProductKitForm, ProductKitItemForm
from .services import search_products
from django.db import transaction
from django.contrib import messages
from django.core.files.base import ContentFile
from openpyxl import load_workbook
import re
from django.core.paginator import Paginator
from django.http import JsonResponse



def is_admin(user):
    if not (user and user.is_active):
        return False
    # allow superusers
    if user.is_superuser:
        return True
    # allow staff
    if getattr(user, 'is_staff', False):
        return True
    # check groups configured in settings.PRODUCTS_ALLOWED_GROUPS
    allowed = getattr(settings, 'PRODUCTS_ALLOWED_GROUPS', ['Admin', 'Manager'])
    try:
        user_groups = set(user.groups.values_list('name', flat=True))
    except Exception:
        user_groups = set()
    return bool(user_groups.intersection(set(allowed)))


def product_list(request):
    qs = Product.objects.select_related('category').all().order_by('code')
    return render(request, 'products/list.html', {'products': qs})


@user_passes_test(is_admin)
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'products/form.html', {'form': form})


def product_detail(request, pk):
    p = get_object_or_404(Product.objects.select_related('category'), pk=pk)
    return render(request, 'products/detail.html', {'product': p})


@user_passes_test(is_admin)
def product_edit(request, pk):
    p = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=p)
        if form.is_valid():
            form.save()
            return redirect('product_detail', pk=p.pk)
    else:
        form = ProductForm(instance=p)
    return render(request, 'products/form.html', {'form': form, 'product': p})


def product_autocomplete(request):
    q = request.GET.get('q', '')
    results = []
    if q:
        items = search_products(q, limit=25)
        for it in items:
            results.append({'id': it.pk, 'text': f"{it.code} - {it.description}"})
    return JsonResponse({'results': results})


@user_passes_test(is_admin)
def product_import(request):
    # Upload POST creates preview and stores rows in session.
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = []
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        seen_codes = set()
        # code pattern: allow letters, numbers, hyphen, underscore, dot and spaces
        code_re = re.compile(r'^[A-Za-z0-9\-_. ]+$')
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw = dict(zip(headers, row))
            r = {k: (v if v is not None else '') for k, v in raw.items()}
            r['_row_number'] = idx
            r['errors'] = {}

            code = (r.get('code') or '').strip()
            if not code:
                r['errors'].setdefault('code', []).append('Código vacío')
            else:
                code = str(code).upper()
                r['code'] = code
                if code in seen_codes:
                    r['errors'].setdefault('code', []).append('Código duplicado en archivo')
                if not code_re.match(code):
                    r['errors'].setdefault('code', []).append('Código con caracteres inválidos')
                seen_codes.add(code)

            # category required and length
            cat_name = (r.get('category') or '').strip()
            if not cat_name:
                r['errors'].setdefault('category', []).append('Categoría vacía')
            else:
                if len(cat_name) > 150:
                    r['errors'].setdefault('category', []).append('Categoría demasiado larga')
                r['category'] = cat_name

            # description validation
            desc = (r.get('description') or '').strip()
            if not desc:
                r['errors'].setdefault('description', []).append('Descripción vacía')
            else:
                if len(desc) > 255:
                    r['errors'].setdefault('description', []).append('Descripción muy larga')
                r['description'] = desc

            # numeric fields validation and sign/limits
            for fld in ('cost', 'price', 'stock', 'stock_min'):
                val = r.get(fld, '')
                if val in (None, ''):
                    r[fld] = None
                    continue
                try:
                    d = Decimal(str(val))
                    if d < 0:
                        r['errors'].setdefault(fld, []).append('Valor negativo')
                    if abs(d) > Decimal('1000000000'):
                        r['errors'].setdefault(fld, []).append('Valor fuera de rango')
                    r[fld] = d
                except (InvalidOperation, TypeError, ValueError):
                    r['errors'].setdefault(fld, []).append(f'{fld} inválido')

            # ruc validation if present
            if 'ruc' in r:
                ruc = (r.get('ruc') or '').strip()
                if ruc:
                    if not re.match(r'^\d{11}$', str(ruc)):
                        r['errors'].setdefault('ruc', []).append('RUC inválido (debe ser 11 dígitos)')

            # cross_reference and location as strings
            r['cross_reference'] = (r.get('cross_reference') or '')
            r['location'] = (r.get('location') or '')

            rows.append(r)

        # detect existing products in DB
        codes = [r['code'] for r in rows if r.get('code')]
        existing = {p.code: p.pk for p in Product.objects.filter(code__in=codes)}
        for r in rows:
            if r.get('code') and r['code'] in existing:
                r['exists'] = True
            else:
                r['exists'] = False

        # store validated rows in session (serialize Decimals)
        serializable = []
        for r in rows:
            s = dict(r)
            for fld in ('cost', 'price', 'stock', 'stock_min'):
                if isinstance(s.get(fld), Decimal):
                    s[fld] = str(s[fld])
            serializable.append(s)

        request.session['import_rows'] = serializable
        request.session['import_headers'] = headers
        # redirect to preview page (GET) to allow pagination
        return redirect('product_import')
    # GET: show preview from session with pagination
    if request.method == 'GET' and request.session.get('import_rows'):
        headers = request.session.get('import_headers', [])
        all_rows = request.session.get('import_rows', [])
        # restore Decimals as strings already; for display keep strings
        # handle pagination
        try:
            page = int(request.GET.get('page', 1))
        except Exception:
            page = 1
        per_page = 50
        total = len(all_rows)
        start = (page-1)*per_page
        end = start + per_page
        page_rows = all_rows[start:end]
        # convert stored rows back into display format (errors dict preserved)
        rows = []
        for r in page_rows:
            rr = dict(r)
            for fld in ('cost', 'price', 'stock', 'stock_min'):
                # keep as string for display
                rr[fld] = rr.get(fld, '')
            rows.append(rr)
        pagination = {'page': page, 'per_page': per_page, 'total': total}
        selected = list(request.session.get('import_selected', []))
        return render(request, 'products/import_preview.html', {'rows': rows, 'headers': headers, 'pagination': pagination, 'selected': selected})
    return render(request, 'products/import.html')


@user_passes_test(is_admin)
def product_import_confirm(request):
    action = request.POST.get('action', 'add_new')
    rows = request.session.get('import_rows', [])
    created = 0
    updated = 0
    errors = []
    to_create = []
    to_update = []
    # deserialize rows and skip rows with validation errors
    deserialized = []
    for r in rows:
        rr = dict(r)
        # restore decimals
        for fld in ('cost', 'price', 'stock', 'stock_min'):
            if isinstance(rr.get(fld), str):
                try:
                    rr[fld] = Decimal(rr[fld])
                except Exception:
                    rr[fld] = None
        deserialized.append(rr)

    # determine selected rows from preview (if provided) or from session
    selected_raw = request.POST.getlist('selected_rows')
    selected = set()
    if selected_raw:
        try:
            selected = set(int(x) for x in selected_raw)
        except Exception:
            selected = set()
    else:
        selected = set(request.session.get('import_selected', []))

    codes = [r.get('code') for r in deserialized if r.get('code') and (not selected or r.get('_row_number') in selected)]
    existing_map = {p.code: p for p in Product.objects.filter(code__in=codes)}

    for r in deserialized:
        # if user selected specific rows, skip others
        if selected and r.get('_row_number') not in selected:
            continue

        if r.get('errors'):
            errors.append({'row': r.get('_row_number'), 'errors': r.get('errors')})
            continue
        code = (r.get('code') or '').strip().upper()
        desc = r.get('description') or ''
        cat_name = (r.get('category') or '').strip()
        cat = None
        if cat_name:
            cat, _ = Category.objects.get_or_create(name=cat_name)

        if code in existing_map:
            if action == 'update_existing':
                p = existing_map[code]
                p.description = desc
                p.cross_reference = r.get('cross_reference') or ''
                p.category = cat
                if r.get('cost') is not None:
                    p.cost = r.get('cost')
                if r.get('price') is not None:
                    p.price = r.get('price')
                if r.get('stock') is not None:
                    p.stock = r.get('stock')
                p.location = r.get('location') or p.location
                to_update.append(p)
                updated += 1
            elif action == 'ignore_existing':
                continue
        else:
            p = Product(
                code=code,
                description=desc,
                cross_reference=r.get('cross_reference') or '',
                category=cat,
                cost=r.get('cost') or 0,
                price=r.get('price') or 0,
                stock=r.get('stock') or 0,
                location=r.get('location') or '',
            )
            to_create.append(p)
            created += 1

    with transaction.atomic():
        if to_create:
            Product.objects.bulk_create(to_create)
        if to_update:
            Product.objects.bulk_update(to_update, ['description', 'cross_reference', 'category', 'cost', 'price', 'stock', 'location'])
    # clear selection and rows after import
    try:
        del request.session['import_selected']
    except Exception:
        pass
    try:
        del request.session['import_rows']
        del request.session['import_headers']
    except Exception:
        pass

    summary = {'created': created, 'updated': updated, 'errors': errors}
    return render(request, 'products/import_summary.html', {'summary': summary})


@user_passes_test(is_admin)
def upload_images_zip(request):
    summary = {'saved': [], 'not_found': []}
    if request.method == 'POST' and request.FILES.get('zip'):
        f = request.FILES['zip']
        z = zipfile.ZipFile(f)
        for name in z.namelist():
            base = name.rsplit('/', 1)[-1]
            if not base:
                continue
            code = base.rsplit('.', 1)[0].upper()
            try:
                p = Product.objects.get(code=code)
                data = z.read(name)
                p.image.save(base, ContentFile(data), save=True)
                summary['saved'].append(code)
            except Product.DoesNotExist:
                summary['not_found'].append(code)
    return render(request, 'products/images_summary.html', {'summary': summary})


@user_passes_test(is_admin)
def product_import_selection(request):
    # AJAX endpoint to persist selected row numbers across preview pages
    if request.method == 'POST':
        visible = request.POST.getlist('visible[]') or request.POST.getlist('visible')
        selected = request.POST.getlist('selected[]') or request.POST.getlist('selected')
        try:
            visible = [int(x) for x in visible]
        except Exception:
            visible = []
        try:
            selected = [int(x) for x in selected]
        except Exception:
            selected = []
        sess = set(request.session.get('import_selected', []))
        # remove any visible items from sess, then add selected
        for v in visible:
            sess.discard(v)
        for s in selected:
            sess.add(s)
        request.session['import_selected'] = list(sess)
        return JsonResponse({'ok': True, 'selected_count': len(sess)})
    return JsonResponse({'ok': False}, status=400)


@user_passes_test(is_admin)
def kit_list(request):
    qs = ProductKit.objects.all().order_by('name')
    return render(request, 'products/kit_list.html', {'kits': qs})


@user_passes_test(is_admin)
def kit_create(request):
    if request.method == 'POST':
        form = ProductKitForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('kit_list')
    else:
        form = ProductKitForm()
    return render(request, 'products/kit_form.html', {'form': form})


@user_passes_test(is_admin)
def kit_edit(request, pk):
    kit = get_object_or_404(ProductKit, pk=pk)
    if request.method == 'POST':
        form = ProductKitForm(request.POST, instance=kit)
        if form.is_valid():
            form.save()
            return redirect('kit_detail', pk=kit.pk)
    else:
        form = ProductKitForm(instance=kit)
    return render(request, 'products/kit_form.html', {'form': form, 'kit': kit})


def kit_detail(request, pk):
    kit = get_object_or_404(ProductKit, pk=pk)
    items = kit.items.select_related('product').all()
    item_form = ProductKitItemForm()
    return render(request, 'products/kit_detail.html', {'kit': kit, 'items': items, 'item_form': item_form})


@user_passes_test(is_admin)
def kit_add_item(request, pk):
    kit = get_object_or_404(ProductKit, pk=pk)
    if request.method == 'POST':
        form = ProductKitItemForm(request.POST)
        if form.is_valid():
            prod = form.cleaned_data['product']
            qty = form.cleaned_data['quantity']
            existing = ProductKitItem.objects.filter(kit=kit, product=prod).first()
            if existing:
                messages.warning(request, f'El producto {prod.code} ya está en el kit. No se agregó.')
            else:
                ProductKitItem.objects.create(kit=kit, product=prod, quantity=qty)
    return redirect('kit_detail', pk=kit.pk)


@user_passes_test(is_admin)
def kit_remove_item(request, pk, item_pk):
    kit = get_object_or_404(ProductKit, pk=pk)
    item = get_object_or_404(ProductKitItem, pk=item_pk, kit=kit)
    if request.method == 'POST' or request.method == 'GET':
        item.delete()
    return redirect('kit_detail', pk=kit.pk)


@user_passes_test(is_admin)
def category_list(request):
    q = request.GET.get('q', '')
    page = int(request.GET.get('page', 1))
    qs = Category.objects.all().order_by('name')
    if q:
        qs = qs.filter(name__icontains=q)
    paginator = Paginator(qs, 20)
    pag = paginator.get_page(page)
    return render(request, 'products/category_list.html', {'categories': pag, 'paginator': paginator, 'q': q})


@user_passes_test(is_admin)
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            cat = form.save()
            messages.success(request, 'Categoría creada.')
            # If AJAX request, return JSON for dynamic add
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'id': cat.pk, 'name': cat.name})
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'products/category_form.html', {'form': form})


@user_passes_test(is_admin)
def category_edit(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=cat)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=cat)
    return render(request, 'products/category_form.html', {'form': form, 'category': cat})


@user_passes_test(is_admin)
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        reassign_to = request.POST.get('reassign_to')
        if reassign_to:
            try:
                other = Category.objects.get(pk=int(reassign_to))
                Product.objects.filter(category=cat).update(category=other)
            except Exception:
                pass
        else:
            # if no reassignment, set products' category to null
            Product.objects.filter(category=cat).update(category=None)
        cat.delete()
        messages.success(request, 'Categoría eliminada.')
        return redirect('category_list')

    categories = Category.objects.exclude(pk=cat.pk).order_by('name')
    return render(request, 'products/category_confirm_delete.html', {'category': cat, 'categories': categories})
