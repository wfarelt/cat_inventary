from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse

from .models import StockMovement, MovementType
from . import services


def movements_list(request):
    qs = StockMovement.objects.select_related('product', 'created_by').order_by('-created_at')
    product_id = request.GET.get('product')
    mtype = request.GET.get('type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if product_id:
        qs = qs.filter(product_id=product_id)
    if mtype:
        qs = qs.filter(movement_type=mtype)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    paginator = Paginator(qs, 25)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'inventory/movements_list.html', {'page_obj': page_obj})


def movement_create(request):
    from inventory.forms import StockMovementForm
    from products.models import Product

    if request.method == 'POST':
        form = StockMovementForm(request.POST)
        # form is only used for basic validation; we use services for changes
        if form.is_valid():
            cd = form.cleaned_data
            product = cd['product']
            movement_type = cd['movement_type']
            quantity = cd['quantity']
            unit_cost = cd.get('unit_cost')
            reason = cd.get('reason') or 'Adjustment'
            reference = cd.get('reference')
            # disallow creating PURCHASE or SALE here (business rule)
            if movement_type in (MovementType.PURCHASE, MovementType.SALE):
                form.add_error('movement_type', 'No permitido crear movimientos de compra/venta manualmente')
            else:
                try:
                    if movement_type == MovementType.CUSTOMER_RETURN:
                        services.register_return(product, quantity, MovementType.CUSTOMER_RETURN, user=request.user, reason=reason, reference=reference)
                    elif movement_type == MovementType.SUPPLIER_RETURN:
                        services.register_return(product, quantity, MovementType.SUPPLIER_RETURN, user=request.user, reason=reason, reference=reference)
                    elif movement_type in (MovementType.ADJUSTMENT_IN, MovementType.INITIAL_STOCK):
                        services.increase_stock(product, quantity, user=request.user, movement_type=movement_type, unit_cost=unit_cost, reason=reason, reference=reference)
                    elif movement_type == MovementType.ADJUSTMENT_OUT:
                        services.decrease_stock(product, quantity, user=request.user, movement_type=movement_type, unit_cost=unit_cost, reason=reason, reference=reference)
                    elif movement_type == MovementType.MANUAL_CORRECTION:
                        # decide direction via hidden field 'direction' in POST
                        direction = request.POST.get('direction')
                        if direction == 'in':
                            services.increase_stock(product, quantity, user=request.user, movement_type=movement_type, unit_cost=unit_cost, reason=reason, reference=reference)
                        else:
                            services.decrease_stock(product, quantity, user=request.user, movement_type=movement_type, unit_cost=unit_cost, reason=reason, reference=reference)
                    messages.success(request, 'Movimiento registrado')
                    return redirect(reverse('inventory:movements_list'))
                except Exception as e:
                    form.add_error(None, str(e))
    else:
        form = StockMovementForm()
    # remove PURCHASE and SALE from selection in form
    form.fields['movement_type'].choices = [c for c in StockMovement._meta.get_field('movement_type').choices if c[0] not in (MovementType.PURCHASE, MovementType.SALE)]
    return render(request, 'inventory/movement_form.html', {'form': form})


def kardex_view(request):
    # filters and export
    qs = StockMovement.objects.select_related('product', 'created_by').order_by('created_at')
    product_id = request.GET.get('product')
    mtype = request.GET.get('type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if product_id:
        qs = qs.filter(product_id=product_id)
    if mtype:
        qs = qs.filter(movement_type=mtype)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    export = request.GET.get('export')
    if export == 'xlsx':
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Kardex'
            headers = ['Fecha', 'Tipo', 'Producto', 'Cantidad', 'Stock Anterior', 'Stock Nuevo', 'Costo', 'Usuario', 'Referencia', 'Motivo']
            ws.append(headers)
            for m in qs:
                ws.append([
                    m.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    m.movement_type,
                    str(m.product),
                    float(m.quantity),
                    float(m.previous_stock),
                    float(m.new_stock),
                    float(m.unit_cost) if m.unit_cost is not None else '',
                    str(m.created_by) if m.created_by else '',
                    m.reference or '',
                    m.reason,
                ])
            # auto-width
            for i, col in enumerate(ws.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=kardex.xlsx'
            wb.save(response)
            return response
        except ImportError:
            # fallback CSV
            import csv
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=kardex.csv'
            writer = csv.writer(response)
            writer.writerow(['Fecha', 'Tipo', 'Producto', 'Cantidad', 'Stock Anterior', 'Stock Nuevo', 'Costo', 'Usuario', 'Referencia', 'Motivo'])
            for m in qs:
                writer.writerow([m.created_at, m.movement_type, str(m.product), m.quantity, m.previous_stock, m.new_stock, m.unit_cost, str(m.created_by) if m.created_by else '', m.reference or '', m.reason])
            return response

    paginator = Paginator(qs, 50)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'inventory/kardex.html', {'page_obj': page_obj})
